'''
workbook = planilha
sheet = pagina
'''

import openpyxl

workbook = openpyxl.Workbook()
#mostra Sheets existentes
print(workbook.sheetnames)
#criando sheets
workbook.create_sheet('ruas')
workbook.create_sheet('cidades')
workbook.create_sheet('estados')
print(workbook.sheetnames)
workbook.save('endereços.xlsx')
#alterando o nome da sheet
workbook['ruas'].title = 'ruas da cidade'
workbook.save('endereços.xlsx')

del workbook['Sheet']
print(workbook.sheetnames)
workbook.save('endereços.xlsx')
